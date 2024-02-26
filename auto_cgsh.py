import json
import time
import openpyxl
import requests
import math
from configparser import ConfigParser
import xlwings as xw
from time import strftime, localtime
import os

CONFIGFILE = '.\\config\\config_cgsh.ini'
config = ConfigParser()
config.read(CONFIGFILE)

tuo = config['base']['tuo']
send_host = config['base']['send_host']
name = config['base']['name']
url = config['base']['get_url']


def get_sql(cgddh, cgddhh):
    try:
        sql = f'select cgddh,cgddhh,sagegysbm,cgyldh from dbo.cgnxjhdmx where cgddh={cgddh} and cgddhh={cgddhh}'
        wlmjson = requests.post('http://10.10.250.30:5002/getjsonone', data={"sql": sql}).text
        tmptask = json.loads(wlmjson)
        return tmptask['sagegysbm'], tmptask['cgyldh']
    except Exception:
        return '', ''


def s_f(len_str, lens, ws, rows):
    row = math.ceil(len_str/lens)
    if ws.row_dimensions[rows].height >= 20*row:
        return
    else:
        ws.row_dimensions[rows].height = 15*row


def print_file_(filename):
    try:
        app = xw.App(visible=False, add_book=False)
        workbook = app.books.open(filename)
        workbook.api.PrintOut(Copies=1, Collate=True)
        workbook.close()
        app.quit()
        print('开始打印...')
        print('打印成功')
    except Exception as e:
        print(e)
        print('打印出错！')


def main():
    try:
        res = requests.get(url=f'{url}{name}').text
        data = json.loads(res)
        print(data)
    except Exception as e:
        # print(e)
        time.sleep(3)
        return

    if data != None:
        wb = openpyxl.load_workbook('./cgsh.xlsx')
        ws = wb.active
        sheetnum = 0

        s1 = []
        for x in range(len(data)):
            gys_num = data[x]['LIFNR']
            s1.append(gys_num)
        s = []
        for i in s1:
            if i not in s:
                s.append(i)
        sheetnum += len(s)

        LI_names = ''
        for i in s:
            LI_names += '-' + i.lstrip("0")

        data_num = 0
        for tmprow in s:
            for x in range(len(data)):
                if data[x]['LIFNR'] == tmprow:
                    data_num += 1
            if data_num > 6:
                if data_num % 6 == 0:
                    sheetnum += (data_num // 6) - 1
                else:
                    sheetnum += data_num // 6
            data_num = 0
        for y in range(sheetnum - 1):
            newwb = wb.copy_worksheet(ws)
            newwb.title = f'Sheet{y+2}'

        t = 1
        lst1 = []
        for subtmprow in s:
            for x in range(len(data)):
                if data[x]['LIFNR'] == subtmprow:
                    t += 1
            lst1.append(t-1)
            t = 1
        dict1 = dict(zip(s, lst1))

        m = 1
        m1 = 1
        s_num = 0
        n = 1
        n1 = 1
        nx = 1
        lst = []
        lst1 = []
        y_num = 1
        for tmpsubrow in dict1.items():
            j = 0
            j1 = 0
            if (tmpsubrow[1] % 6) == 0:
                Y_num = tmpsubrow[1]//6
            else:
                Y_num = tmpsubrow[1]//6 + 1
            for k in range(len(data)):
                ws = wb[f'Sheet{m}']
                time_end = strftime('%Y-%m-%d %H:%M:%S', localtime())
                ws['I21'].value = time_end
                ws['B21'].value = data[k]['FULLNAME']
                ws['E21'].value = data[k]['FULLNAME']
                ws['B2'].value = data[k]['LGORT'] + '/' + data[k]['LGORT_T']
                ws['I2'].value = f'第{y_num}页，共{Y_num}页'
                tmptask = get_sql(data[k]['EBELN'], data[k]['EBELP'])
                if data[k]['LIFNR'] == tmpsubrow[0]:
                    ws['B3'].value = tmpsubrow[0].lstrip("0") + '/' + data[k]['NAME_ORG1'] + '/' + tmptask[0]
                    lst.append(data[k]['MATNR'].lstrip("0"))
                    if data[k]['BUDAT_MKPF'][0:4] == '9999':
                        if data[k]['PART_NO'] == '' and data[k]['DRAWING_NO'] == '':
                            ws[f'B{j + 6}'].value = ' '
                        elif data[k]['PART_NO'] != '' and data[k]['DRAWING_NO'] == '':
                            ws[f'B{j + 6}'].value = data[k]['PART_NO']
                        elif data[k]['PART_NO'] == '' and data[k]['DRAWING_NO'] != '':
                            ws[f'B{j + 6}'].value = data[k]['DRAWING_NO']
                        else:
                            ws[f'B{j + 6}'].value = data[k]['PART_NO']
                        ws[f'E{j + 6}'].value = data[k]['MATNR'].lstrip("0")
                        ws[f'E{j + 7}'].value = '小计数量：'
                        ws[f'G{j + 7}'].value = data[k]['ZXJ_SL']
                        ws[f'I{j + 7}'].value = '小计金额：'
                        ws[f'K{j + 7}'].value = data[k]['ZXJ_JE']
                        j += 2
                        n1 += 1
                    else:
                        if len(lst) == 1:
                            ws[f'A{j + 6}'].value = nx
                            nx += 1
                        else:
                            if lst[-1] != lst[-2]:
                                ws[f'A{j + 6}'].value = nx
                                nx += 1
                        JT_len = len(data[k]['PART_NO'] + '/' + data[k]['DRAWING_NO'])
                        s_f(JT_len, 40, ws, j + 6)
                        if data[k]['PART_NO'] == '' and data[k]['DRAWING_NO'] == '':
                            ws[f'B{j + 6}'].value = ' '
                        elif data[k]['PART_NO'] != '' and data[k]['DRAWING_NO'] == '':
                            ws[f'B{j + 6}'].value = data[k]['PART_NO']
                        elif data[k]['PART_NO'] == '' and data[k]['DRAWING_NO'] != '':
                            ws[f'B{j + 6}'].value = data[k]['DRAWING_NO']
                        else:
                            ws[f'B{j + 6}'].value = data[k]['PART_NO']
                        ws[f'E{j + 6}'].value = data[k]['MATNR'].lstrip("0")
                        ws[f'H{j + 6}'].value = data[k]['STPRS']
                        s_f(len(data[k]['MAKTX']), 32, ws, j + 7)
                        ws[f'B{j + 7}'].value = data[k]['MAKTX']
                        ws[f'E{j + 7}'].value = data[k]['MBLNR']
                        ws[f'G{j + 7}'].value = data[k]['MENGE']
                        ws[f'H{j + 7}'].value = data[k]['DMBTR']
                        ws[f'I{j + 7}'].value = data[k]['EBELN']
                        ws[f'I{j + 6}'].value = data[k]['CHARG']
                        ws[f'K{j + 6}'].value = data[k]['BUDAT_MKPF']
                        ws[f'K{j + 7}'].value = tmptask[1]
                        if data[k]['SOBKZ'] == '':
                            ws[f'G{j + 6}'].value = '非零'
                        else:
                            ws[f'G{j + 6}'].value = '零'
                        j += 2
                        n += 1
                        n1 += 1
                    if j + 6 > 16:
                        m += 1
                        j -= 12
                        y_num += 1
            MENGE_num = 0
            DMBTRL_num = 0
            DMBTRFL_num = 0
            lst1.append(n1-1)
            n_num1 = lst1[-1]
            for i in range(len(data)):
                ws = wb[f'Sheet{m1}']
                ws['H2'].value = nx-1
                if data[i]['LIFNR'] == tmpsubrow[0]:
                    if data[i]['BUDAT_MKPF'][0:4] != '9999':
                        MENGE_num += float(data[i]['MENGE'])
                        if data[i]['SOBKZ'] == '':
                            DMBTRFL_num += float(data[i]['DMBTR'])
                        else:
                            DMBTRL_num += float(data[i]['DMBTR'])
                        j1 += 2
                    if j1 + 6 > 16:
                        m1 += 1
                        j1 -= 12
                    if n_num1 <= 6:
                        ws['A18'].value = '非零库计划金额：'
                        ws['C18'].value = DMBTRFL_num
                        ws['G18'].value = '零库计划金额：'
                        ws['I18'].value = DMBTRL_num
                        ws['A19'].value = '数量合计：'
                        ws['C19'].value = MENGE_num
                        ws['G19'].value = '金额合计：'
                        ws['I19'].value = DMBTRL_num + DMBTRFL_num
            if n_num1 <= 6:
                s_num += 1
            else:
                if (n_num1 % 6) == 0:
                    s_num += (n_num1 // 6)
                    ws = wb[f'Sheet{s_num}']
                    ws['A18'].value = '非零库计划金额：'
                    ws['C18'].value = DMBTRFL_num
                    ws['G18'].value = '零库计划金额：'
                    ws['I18'].value = DMBTRL_num
                    ws['A19'].value = '数量合计：'
                    ws['C19'].value = MENGE_num
                    ws['G19'].value = '金额合计：'
                    ws['I19'].value = DMBTRL_num + DMBTRFL_num
                else:
                    s_num += (n_num1 // 6) + 1
                    ws = wb[f'Sheet{s_num}']
                    ws['A18'].value = '非零库计划金额：'
                    ws['C18'].value = DMBTRFL_num
                    ws['G18'].value = '零库计划金额：'
                    ws['I18'].value = DMBTRL_num
                    ws['A19'].value = '数量合计：'
                    ws['C19'].value = MENGE_num
                    ws['G19'].value = '金额合计：'
                    ws['I19'].value = DMBTRL_num + DMBTRFL_num
            m += 1
            m1 += 1
            n = 1
            n1 = 1
            y_num = 1
            nx = 1
            if tmpsubrow[1] % 6 == 0:
                m -= 1
                m1 -= 1
        wb.save(f'./采购收货单/cgsh采购收货单{LI_names}.xlsx')
        # os.startfile(f'.\\采购收货单\\cgsh采购收货单{LI_names}.xlsx')
    print(f'采购收货单输出成功(cgsh采购收货单{LI_names})')
    print_file_(f'.\\采购收货单\\cgsh采购收货单{LI_names}.xlsx')


if __name__ == "__main__":
    while True:
        try:
            main()
            time.sleep(1)
        except Exception as E:
            print(E)
